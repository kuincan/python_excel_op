# -*- coding: utf-8 -*-
"""
Created on Wed July  28 18:42:05 2021

@author: zhangsiru
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jun  9 16:45:50 2021

@author: yuanzhu
"""

# -*- coding: utf-8 -*-
"""
Created on Sat May  8 12:08:05 2021

@author: yanyujie
"""


from openpyxl.styles import PatternFill
import openpyxl
import os

#dir0=r"C:/Users/yuanzhu/Desktop/table/multiData"
print("zheng zai huizong")
dir0=os.getcwd()
print(dir0)
excel_dir=dir0+"\\data"

from_path = dir0+'\\result'
from_excel_list = os.listdir(from_path)
from_excel = os.path.join(from_path, from_excel_list[0])
# from_excel=dir0+"\\result\\result.xlsx"
from_excel_name = os.listdir(from_path)[0]
print(from_excel_name)
to_excel_name = from_excel_name.split('.')[0]+'_label.xlsx'
to_excel = os.path.join(from_path, to_excel_name)
print(to_excel)

filelist = os.listdir(excel_dir)
#resultfile = os.listdir(save_dir)

excel1 = os.path.join(excel_dir, filelist[0])
excel2 = os.path.join(excel_dir, filelist[1])
excel3 = os.path.join(excel_dir, filelist[2])


def get_one_row(sheet, min_row, max_row, min_col):
    one_row = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, values_only=True):
        for cell in row:
            if cell is not None:
                if len(cell) < 3:
                    one_row.append(' ')
                else:
                    one_row.append(cell)
    return one_row


def get_one_col(sheet, min_row, min_col, max_col):
    one_col = []
    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, values_only=True):
        # print(col)
        for cell in col:
            if cell is not None:
                if len(cell) < 3:
                    one_col.append(' ')
                else:
                    one_col.append(cell)
    return one_col


def comparedata(excel1, excel2, excel3, from_excel, to_excel):

    wb=openpyxl.load_workbook(from_excel)
    print(from_excel)
    sheetnames = wb.sheetnames  # 返回 ['sheet1', 'sheet2',……]

    # print(sheetnames)

    sheet_count = len(sheetnames)

    excel_1 = openpyxl.load_workbook(excel1)
    excel_2 = openpyxl.load_workbook(excel2)
    excel_3 = openpyxl.load_workbook(excel3)
    print(excel1)
    print(excel2)
    print(excel3)

    # excel_1=xlrd.open_workbook(excel1) #打开excel文件
    # excel_2=xlrd.open_workbook(excel2)
    # excel_3 = xlrd.open_workbook(excel3)


    for i in range(sheet_count):
        # print(sheetnames[i])
        sheet1 = excel_1[sheetnames[i]]

        row_value1 = get_one_row(sheet1, min_row=1, max_row=1, min_col=5)
        col_value1 = get_one_col(sheet1, min_col=1, max_col=1, min_row=5)

        sheet2 = excel_2[sheetnames[i]]

        row_value2 = get_one_row(sheet2, min_row=1, max_row=1, min_col=5)
        col_value2 = get_one_col(sheet2, min_col=1, max_col=1, min_row=5)

        sheet3 = excel_3[sheetnames[i]]
        row_value3 = get_one_row(sheet3, min_row=1, max_row=1, min_col=5)
        col_value3 = get_one_col(sheet3, min_col=1, max_col=1, min_row=5)

        # print(row_value1, col_value1)

        new_col = []
        col_len = max(len(col_value1),len(col_value2), len(col_value3))
        for j in range(col_len):
            # print('col:'+str(j))
            test_col=[]
            if j<len(col_value1):
                 test_col.append(col_value1[j])
            else:
                 test_col.append(" ")

            if j<len(col_value2):
                 test_col.append(col_value2[j])
            else:
                 test_col.append(" ")

            if j < len(col_value3):
                test_col.append(col_value3[j])
            else:
                test_col.append(" ")

            a_dic={}
            for k in test_col:
                a_dic[k]=a_dic.get(k,0)+1 # 统计每一个标签出现的次数
            a_dic=sorted(a_dic.items(), key=lambda item:item[1] ,reverse=True)
            a_dic=a_dic[0]
            # print(a_dic)
            # print(a_dic[0]) # 标签
            # print(a_dic[1]) # 出现次数

            if a_dic[1] >= 3:  # 有三人+相同
                new_col.append(a_dic[0])
            else:
                new_col.append(''.join(test_col))

        new_row=[]
        row_len=max(len(row_value1),len(row_value2), len(row_value3))
        # print(sheetnames[i])
        for r in range(row_len):
            # print('row:'+str(r))
            test_row=[]
            if r<len(row_value1):
                 test_row.append(row_value1[r])
            else:
                 test_row.append(" ")

            if r<len(row_value2):
                 test_row.append(row_value2[r])
            else:
                 test_row.append(" ")

            if r < len(row_value3):
                test_row.append(row_value3[r])
            else:
                test_row.append(" ")

            b_dic={}  # 统计每个单元格相同标签个数
            # print(test_row)
            for k in test_row:
                # print(k)
                b_dic[k]=b_dic.get(k,0)+1
            b_dic=sorted(b_dic.items(), key=lambda item:item[1] ,reverse=True)
            b_dic=b_dic[0]


            if b_dic[1] >= 3:
                new_row.append(b_dic[0])
            else:
                new_row.append(''.join(test_row))

        sheet=wb[sheetnames[i]]
        orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")
        c_j=1
        c_i=5
        flag = 0
        col_count = 0
        for datas in new_col:
            sheet.cell(row=c_i, column=c_j).value = datas
            if len(datas) > 3:
                sheet.cell(row=c_i, column=c_j).fill = orange_fill
                flag = 1  # 有不同标签

            c_i=c_i+1
            col_count = col_count+1

        r_i=1
        r_j=5
        row_count = 0
        for datas in new_row:
            sheet.cell(row=r_i, column=r_j).value = datas
            if len(datas) > 3:
                sheet.cell(row=r_i, column=r_j).fill = orange_fill
                flag = 1  # 有不同标签

            row_count = row_count+1

            r_j=r_j+1

        if flag == 1:
            sheet.sheet_properties.tabColor = '7030A0'
        # print(sheetnames[i])
    wb.save(to_excel)


comparedata(excel1,excel2, excel3, from_excel,to_excel)
print("done")
