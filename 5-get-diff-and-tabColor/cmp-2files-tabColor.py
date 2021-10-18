# -*- coding: utf-8 -*-
# @Author  : Zhangsiru
# @Time    : 2021/8/2 10:07
# @Function: 针对筛表任务，获取含标黄标签的文件

import openpyxl
import os

dir0 = os.getcwd()
file_dir = dir0+'\\after'
files_list = os.listdir(file_dir)

raw_dir = dir0+'\\before'
raws_list = os.listdir(raw_dir)

result_txt_path = os.path.join(dir0, 'result_cmp_2files.txt')
# print(result_txt)
result_file = open(result_txt_path, mode='w')


def get_diff(file_list, raws_list):
    for i in range(len(file_list)):
        if i >= len(file_list) or i >= len(raws_list):
            break
        print(file_list[i])
        print(raws_list[i])

        flag = 0
        flag2 = 0
        count_diff = 0
        file_path = os.path.join(file_dir, file_list[i])
        raw_path = os.path.join(raw_dir, raws_list[i])

        wb_file = openpyxl.load_workbook(file_path)
        wb_raw = openpyxl.load_workbook(raw_path)

        sheetnames = wb_file.sheetnames
        for j in range(len(sheetnames)):
            sheet_file = wb_file[sheetnames[j]]
            sheet_raw = wb_raw[sheetnames[j]]
            sheet_file_color = sheet_file.sheet_properties.tabColor.rgb
            sheet_raw_color = sheet_raw.sheet_properties.tabColor.rgb
            # print(sheet_file_color)
            # print(sheet.sheet_properties.tabColor.rgb)
            # if wsprops.tabColor.rgb == 'FFFFFF00':  # 因为openpyxl要求aRGB的数据，excle中显示的是RGB，需要自己转化一下
            if sheet_file_color != sheet_raw_color:
                flag = 1  # 有改动
                count_diff = count_diff+1
                print(sheet_file_color)

            if sheet_file_color == 'FFFF0000' or sheet_file_color == 'FF008000':
                flag2 = 2

            # 标签有不同flag为1，标签非红绿 flag2为2
            # 1. flag == 1
            # 2. flag2 == 2
            # 1 2
            # 1 0
            # 0 2 不可能
            # 0 0 直接过

            if flag == 1:  # 1 0
                # sheet_file.sheet_properties.tabColor = '7030A0'
                result_file.write(str(file_list[i]) + ':' + str(sheetnames[j]))
                if flag2 != 2:
                    result_file.write('  error')
                result_file.write('\n')

            # print(flag)

        if flag == 1:
            result_file.write(str(file_list[i]) + ':' + str(count_diff) + '\n')
            result_file.write('\n======================================================\n')

        wb_file.save(file_path)
        wb_file.close()
        wb_raw.close()


get_diff(file_list=files_list, raws_list=raws_list)
print('done')
result_file.close()
