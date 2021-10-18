# -*- coding: utf-8 -*-
# @Author  : Zhangsiru
# @Time    : 2021/8/2 10:07
# @Function: 针对筛表任务，获取含标黄标签的文件

import openpyxl
import os

dir0 = os.getcwd()
print(dir0)
file_dir = dir0+'\\after'
files_list = os.listdir(file_dir)

result_txt = os.path.join(dir0, 'result.txt')
# print(result_txt)
result_file = open(result_txt, mode='w')


def get_diff(file_list):
    for i in range(len(file_list)):
        print(file_list[i])
        flag = 0
        count_diff = 0
        file_path = os.path.join(file_dir, file_list[i])
        wb = openpyxl.load_workbook(file_path)
        sheetnames = wb.sheetnames
        for j in range(len(sheetnames)):
            sheet = wb[sheetnames[j]]
            # print(sheet.sheet_properties.tabColor.rgb)
            # if wsprops.tabColor == 'FFFFFF00':  # 因为openpyxl要求aRGB的数据，excle中显示的是RGB，需要自己转化一下
            # if sheet.sheet_properties.tabColor.rgb == 'FFFFFF00':  # 黄色
            # print(sheet.sheet_properties.tabColor.rgb)

            if sheet.sheet_properties.tabColor.rgb == '007030A0':  # 紫色

                flag = 1
                count_diff = count_diff+1
            # print(flag)

        if flag == 1:
            result_file.write(str(file_list[i]) + ':' + str(count_diff) + '\n')

        wb.close()


get_diff(files_list)
print('done')
result_file.close()
