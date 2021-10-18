# -*- coding: utf-8 -*-
# @Author  : Zhangsiru
# @Time    : 2021/8/12 0:42
# @Function: compare result and raw file, and color the different sheets and cells.

import openpyxl
from openpyxl.styles import Font
import os
import operator

dir0 = os.getcwd()
print(dir0)
result_file_path = dir0+'\\result'  # 该文件夹下可以有多个待比较的文件
result_file_list = os.listdir(result_file_path)

raw_file_path = dir0+'\\raw'  # 该文件夹下有多个原始文件
raw_file_list = os.listdir(raw_file_path)

outcome_path = dir0+'\\color_result'

cmp_txt_path = os.path.join(dir0, 'cmp_result.txt')
cmp_txt_file = open(cmp_txt_path, mode='w')

font_red = Font(color='FFFF0000')


def get_one_row(sheet, min_row, max_row, min_col):
    one_row = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, values_only=True):
        for cell in row:
            if cell is not None:
                if len(cell) < 3:
                    one_row.append(' ')
                else:
                    one_row.append(cell)
    return one_row


def get_one_col(sheet, min_row, min_col, max_col):
    one_col = []
    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, values_only=True):
        # print(col)
        for cell in col:
            if cell is not None:
                if len(cell) < 3:
                    one_col.append(' ')
                else:
                    one_col.append(cell)
    return one_col


def cmp_files_color(raw_list, result_list):
    """
    :param raw_list: 多个原始文件的文件名列表
    :param result_list: 多个结果文件的文件名列表
    :return:
    """
    # 以下操作均以result文件为基准

    count_color_sheet_all = 0

    for i in range(len(result_list)):
        if i >= len(raw_list):
            print('Need more raw files.\n')
            exit(-1)
        print(raw_list[i])
        print(result_list[i])

        count_color_sheet = 0

        raw_path = os.path.join(raw_file_path, raw_list[i])
        result_path = os.path.join(result_file_path, result_list[i])

        wb_raw = openpyxl.load_workbook(raw_path)
        wb_result = openpyxl.load_workbook(result_path)

        color_result_path = os.path.join(outcome_path, result_list[i])  # 输出文件路径

        sheet_names = wb_result.sheetnames
        for j in range(len(sheet_names)):
            flag_sheet = 0  # color sheet

            sheet_result = wb_result[sheet_names[j]]
            row_result = get_one_row(sheet_result, min_row=1, max_row=1, min_col=5)
            col_result = get_one_col(sheet_result, min_col=1, max_col=1, min_row=5)

            if sheet_names[j] not in wb_raw.sheetnames:
                # color sheet directly
                # color all cells red
                # 从E1开始，行着色
                for idx in range(row_result):
                    sheet_result.cell(row=1, column=5+idx).font = font_red

                # 从A5开始，列着色
                for idx in range(col_result):
                    sheet_result.cell(row=5+idx, column=1).font = font_red

                sheet_result.sheet_properties.tabColor = 'FF0000'  # red
                print(sheet_result.sheet_properties.tabColor)
                flag_sheet = 1
                count_color_sheet = count_color_sheet + 1
                cmp_txt_file.write(str(result_list[i])+':red '+ str(sheet_names[j]) +'\n')
                if j == len(row_result)-1:
                    count_color_sheet_all = count_color_sheet_all+count_color_sheet

                continue
            else:
                sheet_raw = wb_raw[sheet_names[j]]
                # print(sheet_raw.sheet_properties.tabColor.rgb)
                # print(sheet_raw.sheet_properties.tabColor)
                if sheet_raw.sheet_properties.tabColor is not None:
                    sheet_result.sheet_properties.tabColor = 'FF0000'
                    flag_sheet = 1
                    count_color_sheet = count_color_sheet+1
                row_raw = get_one_row(sheet_raw, min_row=1, max_row=1, min_col=5)
                col_raw = get_one_col(sheet_raw, min_col=1, max_col=1, min_row=5)

                if operator.eq(row_result, row_raw) and operator.eq(col_result, col_raw):
                    continue

                # 先看第一行
                for row_idx in range(len(row_result)):
                    if row_idx >= len(row_raw):
                        # 则该行之后所有单元格都要标红
                        for idx_r in range(len(row_result)-row_idx):  # 比如 idx_t=0, row_idx=2, len(row_raw)=2,len(row_result)=4
                            sheet_result.cell(row=1, column=row_idx+idx_r+5).font = font_red
                        flag_sheet = 1
                        continue
                    if row_result[row_idx] != row_raw[row_idx]:
                        sheet_result.cell(row=1, column=row_idx+5).font = font_red
                        flag_sheet = 1

                # 再看第一列
                for col_idx in range(len(col_result)):
                    if col_idx >= len(col_raw):
                        for idx_c in range(len(col_result)-col_idx):
                            sheet_result.cell(row=col_idx+idx_c+5, column=1).font = font_red
                        flag_sheet = 1
                        continue
                    if col_result[col_idx] != col_raw[col_idx]:
                        sheet_result.cell(row=col_idx+5, column=1).font = font_red
                        flag_sheet = 1

            if flag_sheet == 1:
                sheet_result.sheet_properties.tabColor = 'FF0000'  # red
                count_color_sheet = count_color_sheet+1

                cmp_txt_file.write(str(result_list[i])+':red '+str(sheet_names[j])+'\n')

        count_color_sheet_all = count_color_sheet_all + count_color_sheet

        wb_result.save(color_result_path)
        wb_result.close()
        wb_raw.close()
    cmp_txt_file.write('\n========================================\n'+'total: '+str(count_color_sheet_all))


cmp_files_color(raw_list=raw_file_list, result_list=result_file_list)
print('done')
cmp_txt_file.close()
