# -*- coding: utf-8 -*-
"""
Created on Sat July 31 21:50:30 2021

Compare Each File with the Right One, and Get the Accuracy.

@author: zhangsiru
"""

import openpyxl
from openpyxl.styles import PatternFill
import os
import operator

dir0 = os.getcwd()
print(dir0)
file_cmp_path = dir0+'\\cmp'  # 该文件夹下可以有多个待比较的文件
files_cmp = os.listdir(file_cmp_path)

right_file_path = dir0+'\\result'  # 该文件夹下只有一个正确的结果文件
right_file_list = os.listdir(right_file_path)
right_file = os.path.join(right_file_path, right_file_list[0])

cmp_result_path = dir0+'\\cmp_result'


result_txt = os.path.join(dir0, right_file_list[0].split('.')[0]+'.txt')
# print(result_txt)
acc_file = open(result_txt, mode='w')

skip_txt = os.path.join(dir0, right_file_list[0].split('.')[0]+'_skip'+'.txt')
skip_file = open(skip_txt, mode='w')

complicated_txt = os.path.join(dir0, right_file_list[0].split('.')[0]+'_complicated'+'.txt')
complicated_file = open(complicated_txt, mode='w')

blank_txt = os.path.join(dir0, right_file_list[0].split('.')[0]+'_blank'+'.txt')
blank_file = open(blank_txt, mode='w')

purple_fill = PatternFill(fill_type='solid', fgColor='6600CC')  # for cell


def get_one_row(sheet, min_row, max_row, min_col):
    one_row = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, values_only=True):
        for cell in row:
            if cell is not None:
                if len(cell) < 3:
                    one_row.append(' ')
                else:
                    one_row.append(cell)
    return one_row


def get_one_col(sheet, min_row, min_col, max_col):
    one_col = []
    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, values_only=True):
        for cell in col:
            if cell is not None:
                if len(cell) < 3:
                    one_col.append(' ')
                else:
                    one_col.append(cell)
    return one_col


def get_acc(right_excel, waiting_cmp_list):
    """
    :param right_excel: 正确的结果文件
    :param waiting_cmp_list: 待比较的文件列表
    :return: None
    """

    wb_right = openpyxl.load_workbook(right_excel)
    sheet_names_right = wb_right.sheetnames

    skip_sheet_count = 0
    complicated_sheet_count = 0

    sheet_count_right = len(sheet_names_right)

    to_cmp_num = len(waiting_cmp_list)
    print(right_excel)

    only_print_once = False

    for i in range(to_cmp_num):
        print(files_cmp[i])
        file_cmp = os.path.join(file_cmp_path, files_cmp[i])
        result_cmp = os.path.join(cmp_result_path, files_cmp[i])  # 经过高亮的原始文件路径

        wb_cmp = openpyxl.load_workbook(file_cmp)
        sheet_count_cmp = len(wb_cmp.sheetnames)

        false_sheet = 0
        acc_file.write(files_cmp[i]+':'+'\n')

        blank_sheet_count = 0
        blank_file.write(files_cmp[i]+':'+'\n')

        for j in range(sheet_count_right):
            # 因为结果文件存在删表情况，所以sheet_count可能不一致，需要通过sheetname找到对应的表进行结果比较
            name = sheet_names_right[j]
            if '(' in name:
                # print(name)
                if i == 0:
                    skip_sheet_count = skip_sheet_count+1
                    skip_file.write(str(name)+'\n')
                if j == sheet_count_right - 1:
                    skip_file.write('total skip: ' + str(skip_sheet_count))
                continue
            if j == sheet_count_right - 1 and i == 0:  # 以防最后一个sheet也是invalid或empty
                skip_file.write('total skip: '+str(skip_sheet_count))

            sheet_right = wb_right[name]
            # row_right = [cell for row in sheet_right.iter_rows(min_row=1, max_row=1, min_col=5, values_only=True)
            #                       for cell in row]

            # A1的索引是1,1
            row_right = get_one_row(sheet_right, min_row=1, max_row=1, min_col=5)
            col_right = get_one_col(sheet_right, min_col=1, max_col=1, min_row=5)

            if name not in wb_cmp.sheetnames:  # 排除invalid和empty等情况
                continue
            sheet_cmp = wb_cmp[name]
            row_cmp = get_one_row(sheet_cmp, min_row=1, max_row=1, min_col=5)
            col_cmp = get_one_col(sheet_cmp, min_col=1, max_col=1, min_row=5)

            if operator.eq(row_right, row_cmp) and operator.eq(col_right, col_cmp):
                continue
            else:
                # print('\nrow result:')
                # print(row_cmp)
                # print(row_right)
                #
                # print('col result: ')
                # print(col_cmp)
                # print(col_right)
                # print(sheet_right.sheet_properties.tabColor)
                if (len(row_right)+len(col_right)) - (len(row_cmp)+len(col_cmp)) > 3:
                    blank_sheet_count = blank_sheet_count+1
                    blank_file.write(str(name)+'\n')

                if sheet_right.sheet_properties.tabColor is not None:
                    if sheet_right.sheet_properties.tabColor.rgb == 'FF002060':  #深蓝色
                        if i == 0:
                            complicated_sheet_count = complicated_sheet_count+1
                            complicated_file.write(str(name)+'\n')
                            only_print_once = True
                        false_sheet = false_sheet - 1  # 后面再加一次，抵消了

                false_sheet = false_sheet + 1
                sheet_cmp.sheet_properties.tabColor = '6600CC'
                acc_file.write(name + ', ')

                # 因为有人漏标，所以需要补充cmp的list
                if len(row_cmp) != len(row_right):
                    for add_idx in range(len(row_right) - len(row_cmp)):
                        row_cmp.append(None)
                if len(col_cmp) != len(col_right):
                    for add_idx in range(len(col_right) - len(col_cmp)):
                        col_cmp.append(None)

                for row_index in range(len(row_right)):
                    if row_right[row_index] != row_cmp[row_index]:
                        sheet_cmp.cell(row=1, column=row_index + 5).fill = purple_fill
                for col_index in range(len(col_right)):
                    if col_right[col_index] != col_cmp[col_index]:
                        sheet_cmp.cell(row=col_index + 5, column=1).fill = purple_fill

        if only_print_once:
            complicated_file.write('complicated sheets:' + str(complicated_sheet_count))
            only_print_once = False

        wb_cmp.save(result_cmp)
        wb_cmp.close()
        blank_file.write('blank sheets(blank cell over 3): '+str(blank_sheet_count)+'\n\n')

        acc = '{:.2%}'.format((sheet_count_right - false_sheet - skip_sheet_count)/(sheet_count_cmp - skip_sheet_count))
        acc_file.write('\n'+'Wrong sheets: '+str(false_sheet))
        acc_file.write('\n' + 'Accuracy: ' + acc)
        acc_file.write('\n' + 'Total Valid sheets: ' + str(sheet_count_right-skip_sheet_count))
        acc_file.write('\n=================================================='+'\n'+'\n')


get_acc(right_excel=right_file, waiting_cmp_list=files_cmp)
print('done')
